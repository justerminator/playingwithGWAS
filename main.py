import openpyxl
from collections import defaultdict
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell


path = "edited-BSB273CC17HSLGA15_output (1).xlsx"



# open workbook
wb_obj = openpyxl.load_workbook(path)

# get active sheet object
sheet_obj = wb_obj.active

# range of animal categories
row = 4
column = 18
interval = 180

listAnimals = []
for i in range(2, 18):
    listAnimals.append(sheet_obj.cell(row, i).value)

# sort active range
xValue = 18
colStart = 15
colEnd = 336

x = 15


dictDur = defaultdict(list)
anCounter = 0

numBursts = 2

# animal for each column
for animal in range(2, xValue):

    allBursts = []
    i = 15
    while i != colEnd:
        if sheet_obj.cell(i, animal).value == 0:
            break
        # initialize start time
        oneBurst = []
        ptr1 = sheet_obj.cell(i, animal).value
        limit = ptr1 + interval

        j = i
        while j != colEnd and sheet_obj.cell(j, animal).value <= limit:
            if sheet_obj.cell(j, animal).value == 0:
                break
            oneBurst.append((sheet_obj.cell(j, 1).value, sheet_obj.cell(j, animal).value))
            j += 1
        if len(oneBurst) > numBursts:
            allBursts.append(oneBurst)
        i = j



    dictDur[listAnimals[anCounter]] = allBursts.copy()
    anCounter += 1
    x = 15



print(dictDur)

newSheet = openpyxl.Workbook()
sheet = newSheet.active

# create title
r = 0
for i in range(len(listAnimals)):
    r += 1
    c1 = sheet.cell(1, r + i + 1)

    c1.value = listAnimals[i]

# populate data
rowCoord = 2
colCoord = 2

grayfill = PatternFill(start_color='00808080', end_color='00808080', fill_type='solid')
for x in range(len(dictDur)):

    for y in range(len(dictDur[listAnimals[x]])):

        for z in range(len(dictDur[listAnimals[x]][y])):

            c2 = sheet.cell(rowCoord, colCoord)

            c2.value = dictDur[listAnimals[x]][y][z][0]

            c3 = sheet.cell(rowCoord, colCoord + 1)
            c3.value = dictDur[listAnimals[x]][y][z][1]

            rowCoord += 1

        c4 = sheet.cell(rowCoord, colCoord)
        c4.fill = grayfill

        c5 = sheet.cell(rowCoord, colCoord + 1)
        c5.fill = grayfill

        rowCoord += 1


    rowCoord = 2
    colCoord += 2

newSheet.save("newfile.xlsx")
