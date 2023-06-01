import csv
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell import Cell


import pandas as pd
import ast
import openpyxl

data = pd.read_excel("xlFull_burst_all_long.xlsx")


data_oxy = data[data["drug_group"] == 'Oxycodone']
iir_data_col = data_oxy[["inter_infusion_interval_LGA12", "rat"]]
iir_data_col = iir_data_col.dropna()


set_intervals = [(20, 39), (40, 59), (60, 79), (80, 99), (100, 119), (120, 139)]
#set_intervals2 = [(20, 29), (40, 59), (60, 79), (80, 99), (100, 119), (120, 139)]

set_intervals2 = []
curr = 0

while curr + 19 != 719:
    curr += 20
    set_intervals2.append((curr, curr + 19))
    print(curr)


wb = openpyxl.Workbook()
ws = wb.active
#sheet = wb['Sheet1']

#TODO: Look at LGA 03, or earlier cohorts

newWb = openpyxl.Workbook()
newSheet = newWb.active

# for x in range(1, len(set_intervals) + 1):
#
#     c1 = ws.cell(1, x + 1)
#     c1.value = str(set_intervals[x - 1])

rowTracker = 2

mapSum = defaultdict(int)

for x in range(len(set_intervals2)):
    mapSum[str(set_intervals2[x])] = 0

def createCounts(lst, rat):

    global rowTracker

    map_counts = defaultdict(int)

    for x in range(len(set_intervals2)):
        map_counts[str(set_intervals2[x])] = 0

    intermed = ast.literal_eval(lst)
    for val in intermed:
        for ind in range(len(set_intervals2)):

            if set_intervals2[ind][0] <= val <= set_intervals2[ind][1]:

                map_counts[str(set_intervals2[ind])] += 1
                mapSum[str(set_intervals2[ind])] += 1
                #print(map_counts)



    # writing to file

    for i in range(0, len(set_intervals2)):
        c2 = ws.cell(rowTracker, 1)
        c2.value = str(rat)

        c3 = ws.cell(rowTracker, 2)
        c3.value = int(i + 1)

        c4 = ws.cell(rowTracker, 3)
        c4.value = int(map_counts[str(set_intervals2[i])])

        rowTracker += 1


    # for i in range(0, 6):
    #     c2 = ws.cell(rowTracker, 1)
    #     c2.value = str(rat)
    #
    #     c2 = ws.cell(rowTracker, i + 2)
    #
    #     print(rat)
    #
    #     c2.value = map_counts[str(set_intervals[i])]



    #rowTracker += 1





    return map_counts



for index, row in iir_data_col.iterrows():

    leRat = row["rat"]
    leLst = row["inter_infusion_interval_LGA12"]


    map = createCounts(leLst, leRat)

for index in range(0, len(set_intervals2)):
    c5 = newSheet.cell(index + 1, 1)
    c5.value = str(set_intervals2[index])

    c6 = newSheet.cell(index + 1, 2)
    c6.value = mapSum[str(set_intervals2[index])]

newWb.save("new.xlsx")
wb.save('newfile2.xlsx')
