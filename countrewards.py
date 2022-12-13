import openpyxl
import main
from collections import defaultdict
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell


path = "120s-BSB273DC10HSOXYLGA12_output (1).xlsx"
grayfill = PatternFill(start_color='00808080', end_color='00808080', fill_type='solid')
yellowfill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')


# open workbook
wb_obj = openpyxl.load_workbook(path)

# get active sheet object
sheet_obj = wb_obj.active

# range of animal categories


def animalArray():
    r = 1
    c = 2
    animalArray = []
    while sheet_obj.cell(r, c).value != None:
        animalArray.append(sheet_obj.cell(r, c).value)
        c += 2
    return animalArray

listAnimals = animalArray()
print(listAnimals)

def countRewards():
    rewardTracker = defaultdict(int)
    row = 2
    column = 3

    i = 0
    while i != len(listAnimals):
        prev = sheet_obj.cell(row, column).value
        sheet_obj.cell(row, column).fill = yellowfill
        rewardTracker[listAnimals[i]] += 1

        total = prev
        #print(prev)
        if prev == None:
            break

        while True:
            row += 1

            next = sheet_obj.cell(row, column).value
            if next == None:
                next = 0

            total = next - prev
            #print(total)
            if total >= 20:
                sheet_obj.cell(row, column).fill = yellowfill
                rewardTracker[listAnimals[i]] += 1
                prev = next


            if next == 0 and sheet_obj.cell(row + 1, column).value == None:
                break

        column += 2
        row = 2
        i += 1
    return rewardTracker

def generateNewSheet(animalLst, workbook):
    rewardData = countRewards()
    workbook.create_sheet('rewards')
    sheet2 = workbook['rewards']

    row = 1
    col = 2
    sheet2.cell(row, col).value = 'rewards'
    for animal in animalLst:
        dataPtr = rewardData[animal]
        sheet2.cell(row, col).value = animal
        sheet2.cell(row + 1, col).value = dataPtr
        sheet2.cell(row + 2, col).value = main.getTotalBursts(animal)
        row = 1
        col += 1
    workbook.save(path)


generateNewSheet(listAnimals, wb_obj)
#print(rewardTracker)
wb_obj.save(path)