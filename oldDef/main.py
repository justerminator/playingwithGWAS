import openpyxl
from collections import defaultdict
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell


path = "Book2.xlsx"
grayfill = PatternFill(start_color='00808080', end_color='00808080', fill_type='solid')
yellowfill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
dictDur = defaultdict(list)


# open workbook
wb_obj = openpyxl.load_workbook(path)

# get active sheet object
sheet_obj = wb_obj.active

# range of animal categories
row = 1
interval = 120
numAnimals = 46 + 2

def getTotalBursts(animal) -> int:
    getBursts()

    return len(dictDur[animal])

def populateAnimals():
    listAnimals = []
    for i in range(2, numAnimals):
        listAnimals.append(sheet_obj.cell(row, i).value)

    return listAnimals

def getBursts():
    # sort active range
    colEnd = 202


    anCounter = 0

    numBursts = 2

    # animal for each column
    for animal in range(2, numAnimals):

        allBursts = []
        i = 2
        while i != colEnd:
            if sheet_obj.cell(i, animal).value == 0 and sheet_obj.cell(i + 1, animal).value == 0:
                break
            # initialize start time
            oneBurst = []
            ptr1 = sheet_obj.cell(i, animal).value
            limit = ptr1 + interval

            j = i
            while j != colEnd and sheet_obj.cell(j, animal).value <= limit:
                if sheet_obj.cell(j, animal).value == 0 and sheet_obj.cell(j + 1, animal).value == 0:
                    break
                oneBurst.append((sheet_obj.cell(j, 1).value, sheet_obj.cell(j, animal).value))
                j += 1
            if len(oneBurst) >= numBursts:
                allBursts.append(oneBurst)
            i = j



        dictDur[listAnimals[anCounter]] = allBursts.copy()
        anCounter += 1



def createNewSheet():
    newSheet = openpyxl.Workbook()
    sheet = newSheet.active
    # create title
    r = 0
    for i in range(len(listAnimals)):

        if dictDur[listAnimals[i]] == []:
            r -= 1
            continue
        r += 1
        c1 = sheet.cell(1, r + i + 1)

        c1.value = listAnimals[i]

    # populate data
    rowCoord = 2
    colCoord = 2
    for x in range(len(dictDur)):
        if dictDur[listAnimals[x]] == []:
            #rowCoord = 2
            #colCoord += 2
            continue

        for y in range(len(dictDur[listAnimals[x]])):

            for z in range(len(dictDur[listAnimals[x]][y])):

                c2 = sheet.cell(rowCoord, colCoord)

                c2.value = dictDur[listAnimals[x]][y][z][0]

                c3 = sheet.cell(rowCoord, colCoord + 1)
                c3.value = dictDur[listAnimals[x]][y][z][1]


                rowCoord += 1

            c4 = sheet.cell(rowCoord, colCoord)
            c4.fill = grayfill

            c5 = sheet.cell(rowCoord, colCoord + 1)
            c5.fill = grayfill

            rowCoord += 1


        rowCoord = 2
        colCoord += 2

    newSheet.save("newBook2.xlsx")

listAnimals = populateAnimals()
def countRewards(data):
    listAnimals = populateAnimals()
    rewardCounter = defaultdict(int)
    rewards = 0
    averageRewards = 0

    for animal in listAnimals:
        listBursts = data[animal]
        for burst in listBursts:
            ptr = None
            for x in range(len(burst)):
                if x != 0 and burst[x][1] - burst[x - 1][1] > 20:
                    if ptr == None:
                        ptr = burst[x][1]
                    rewards += 1
            #averageRewards += rewards / len(burst)
            #rewards = 0
            #print(averageRewards)
        averageRewards = 0
    return rewardCounter


getBursts()
createNewSheet()
#countRewards(dictDur)
print(dictDur)